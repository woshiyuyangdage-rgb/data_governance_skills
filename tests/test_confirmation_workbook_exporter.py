"""Tests for confirmation workbook export."""

from pathlib import Path

from openpyxl import load_workbook

from app.core.delivery.confirmation_workbook_exporter import (
    ConfirmationWorkbookExporter,
)
from app.core.models.governance_backlog_item import GovernanceBacklogItem
from app.core.models.mapping_result import MappingResult
from app.core.models.quality_rule_suggestion import QualityRuleSuggestion
from app.core.models.stg_field_suggestion import StgFieldSuggestion


def test_confirmation_workbook_exporter_exports_four_workbooks(tmp_path: Path) -> None:
    exporter = ConfirmationWorkbookExporter()

    mapping_result = exporter.export_mapping_confirmation_workbook(
        [
            MappingResult(
                table_name="customer",
                field_name="customer_id",
                recommended_standard_code="customer_id",
                match_score=1.0,
                match_reason="exact match",
            )
        ],
        str(tmp_path / "mapping.xlsx"),
    )
    stg_result = exporter.export_stg_confirmation_workbook(
        [
            StgFieldSuggestion(
                source_table_name="customer",
                source_field_name="customer_id",
                recommended_stg_field_name="customer_id",
                mapping_source="standard_mapping",
                action="keep",
            )
        ],
        str(tmp_path / "stg.xlsx"),
    )
    quality_result = exporter.export_quality_rule_confirmation_workbook(
        [
            QualityRuleSuggestion(
                source_table_name="customer",
                source_field_name="customer_id",
                rule_type="not_null",
                rule_expression="not_null",
                severity="high",
                recommendation_source="template",
            )
        ],
        str(tmp_path / "quality.xlsx"),
    )
    backlog_result = exporter.export_backlog_delivery_workbook(
        [
            GovernanceBacklogItem(
                backlog_id="backlog_1",
                object_type="field",
                object_name="customer.customer_id",
                gap_type="quality_rule_gap",
                action="Confirm quality rule",
                owner_role="data_steward",
                priority="high",
                status="proposed",
            )
        ],
        str(tmp_path / "backlog.xlsx"),
    )

    for result in [mapping_result, stg_result, quality_result, backlog_result]:
        assert result.status == "success"
        assert result.row_count == 1
        workbook = load_workbook(result.output_path)
        assert "instructions" in workbook.sheetnames
        assert "summary" in workbook.sheetnames

