"""Report export tests for the local MVP workflow."""

import json
from pathlib import Path

from openpyxl import load_workbook

from app.core.models.backlog_summary import BacklogSummary
from app.core.models.backlog_sla_status import BacklogSlaStatus
from app.core.models.confirmation_workbook_result import ConfirmationWorkbookResult
from app.core.models.confirmed_quality_rule import ConfirmedQualityRule
from app.core.models.cross_field_quality_rule import CrossFieldQualityRule
from app.core.adapters.execution_package_builder import ExecutionPackageBuilder
from app.core.models.execution_package_export_result import ExecutionPackageExportResult
from app.core.models.governance_backlog_item import GovernanceBacklogItem
from app.core.models.governance_delivery_manifest import GovernanceDeliveryManifest
from app.core.models.governance_delivery_package_result import (
    GovernanceDeliveryPackageResult,
)
from app.core.models.batch_group_result import BatchGroupResult
from app.core.models.confirmation_roundtrip_result import ConfirmationRoundTripResult
from app.core.models.incremental_diff_item import IncrementalDiffItem
from app.core.models.incremental_diff_summary import IncrementalDiffSummary
from app.core.models.workbook_import_summary import WorkbookImportSummary
from app.core.models.governance_gap import GovernanceGap
from app.core.models.governance_portfolio_summary import GovernancePortfolioSummary
from app.core.models.governance_work_package import GovernanceWorkPackage
from app.core.models.mapping_review_record import MappingReviewRecord
from app.core.models.progress_snapshot import ProgressSnapshot
from app.core.models.readiness_score import ReadinessScore
from app.core.models.remediation_action import RemediationAction
from app.core.models.rule_export_result import RuleExportResult
from app.core.models.stg_review_record import StgReviewRecord
from app.core.models.workflow_result import WorkflowResult
from app.core.orchestrator.pipeline_service import (
    run_p0_plus_mapping_from_file,
    run_p0_plus_mapping_plus_stg_from_file,
    run_p0_plus_mapping_plus_stg_plus_quality_from_file,
    run_p0_plus_mapping_plus_stg_with_review_from_file,
)
from app.core.orchestrator.workflow_engine import WorkflowEngine
from app.core.reports.excel_reporter import export_workflow_result_to_excel
from app.core.reports.json_reporter import export_workflow_result_to_json
from app.core.reports.markdown_reporter import export_workflow_result_to_markdown
from app.core.reports.report_service import (
    build_report_base_filename,
    export_all_reports,
)
from app.core.review import override_store

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_METADATA_PATH = PROJECT_ROOT / "app" / "data" / "samples" / "sample_metadata.csv"


def _build_demo_result():
    engine = WorkflowEngine()
    return engine.run_p0_pipeline(engine.build_demo_tables())


def test_export_workflow_result_files_are_generated(tmp_path: Path) -> None:
    result = _build_demo_result()
    json_path = tmp_path / "workflow_result.json"
    markdown_path = tmp_path / "workflow_result.md"
    excel_path = tmp_path / "workflow_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    assert json_path.exists()
    assert markdown_path.exists()
    assert excel_path.exists()

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert json_payload["status"] == "success"
    assert "issues" in json_payload
    assert "tasks" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Project Run Summary" in markdown_content
    assert "# Governance Tasks" in markdown_content

    workbook = load_workbook(excel_path)
    assert "summary" in workbook.sheetnames
    assert "issues" in workbook.sheetnames
    assert "tasks" in workbook.sheetnames
    assert "skill_outputs_overview" in workbook.sheetnames


def test_export_all_reports_returns_existing_paths(tmp_path: Path) -> None:
    result = _build_demo_result()

    report_paths = export_all_reports(result, str(tmp_path), "demo_run")

    assert set(report_paths.keys()) == {"json", "markdown", "excel"}
    for report_path in report_paths.values():
        assert Path(report_path).exists()


def test_build_report_base_filename_uses_profile_report_mode() -> None:
    diagnosis_name = build_report_base_filename(
        profile_name="metadata_diagnosis_only",
        timestamp="20260501_120000",
    )
    confirmed_name = build_report_base_filename(
        profile_name="diagnosis_mapping_stg_with_review",
        timestamp="20260501_120000",
    )

    assert diagnosis_name == "diagnosis_20260501_120000"
    assert confirmed_name == "confirmed_20260501_120000"


def test_reports_include_mapping_outputs_when_present(tmp_path: Path) -> None:
    result = run_p0_plus_mapping_from_file(str(SAMPLE_METADATA_PATH))
    json_path = tmp_path / "mapping_workflow_result.json"
    markdown_path = tmp_path / "mapping_workflow_result.md"
    excel_path = tmp_path / "mapping_workflow_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "field_description_suggestions" in json_payload
    assert "table_semantic_summaries" in json_payload
    assert "mapping_results" in json_payload
    assert "unmapped_fields" in json_payload
    assert "risk_hint" in json_payload["mapping_results"][0]
    assert "action_suggestion" in json_payload["mapping_results"][0]

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Semantic Enrichment Suggestions" in markdown_content
    assert "## Field Description Suggestions" in markdown_content
    assert "# Standard Mapping Recommendations" in markdown_content
    assert "# Unmapped or Low-Confidence Fields" in markdown_content
    assert "risk=" in markdown_content
    assert "action=" in markdown_content

    workbook = load_workbook(excel_path)
    assert "field_descriptions" in workbook.sheetnames
    assert "table_semantic_summary" in workbook.sheetnames
    assert "standard_mapping" in workbook.sheetnames
    assert "unmapped_fields" in workbook.sheetnames
    table_summary_headers = [
        cell.value for cell in next(workbook["table_semantic_summary"].iter_rows(max_row=1))
    ]
    assert "business_object" in table_summary_headers
    assert "business_purpose" in table_summary_headers
    assert "core_fields_joined" in table_summary_headers
    assert "applicable_scenarios_joined" in table_summary_headers
    assert "ai_usage_risks_joined" in table_summary_headers
    assert "recommended_actions_joined" in table_summary_headers
    standard_mapping_headers = [
        cell.value for cell in next(workbook["standard_mapping"].iter_rows(max_row=1))
    ]
    assert "risk_hint" in standard_mapping_headers
    assert "action_suggestion" in standard_mapping_headers


def test_reports_include_stg_outputs_when_present(tmp_path: Path) -> None:
    result = run_p0_plus_mapping_plus_stg_from_file(str(SAMPLE_METADATA_PATH))
    json_path = tmp_path / "stg_workflow_result.json"
    markdown_path = tmp_path / "stg_workflow_result.md"
    excel_path = tmp_path / "stg_workflow_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "stg_suggestions" in json_payload
    assert "stg_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# STG Structure Suggestions" in markdown_content

    workbook = load_workbook(excel_path)
    assert "stg_tables" in workbook.sheetnames
    assert "stg_fields" in workbook.sheetnames


def test_reports_include_quality_outputs_when_present(tmp_path: Path) -> None:
    result = run_p0_plus_mapping_plus_stg_plus_quality_from_file(str(SAMPLE_METADATA_PATH))
    json_path = tmp_path / "quality_workflow_result.json"
    markdown_path = tmp_path / "quality_workflow_result.md"
    excel_path = tmp_path / "quality_workflow_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "quality_rule_suggestions" in json_payload
    assert "quality_rule_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Quality Rule Recommendations" in markdown_content

    workbook = load_workbook(excel_path)
    assert "quality_rules" in workbook.sheetnames
    assert "quality_rule_summary" in workbook.sheetnames
    assert "cross_field_quality_rules" in workbook.sheetnames
    assert "quality_review_queue_summary" in workbook.sheetnames


def test_reports_include_cross_field_and_review_queue_outputs(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="cross-field report test",
        cross_field_quality_rules=[
            CrossFieldQualityRule(
                source_table_name="sales_order",
                field_group=["start_date", "end_date"],
                rule_type="temporal_order",
                rule_expression="start_date <= end_date",
                severity="medium",
                confidence=1.0,
                review_priority="medium_review_priority",
                recommendation_source="cross_field_pattern",
                match_basis="start_date/end_date",
                reason="Start date should not be later than end date.",
            )
        ],
        quality_review_queue_summary={
            "total_rule_count": 1,
            "field_rule_count": 0,
            "cross_field_rule_count": 1,
            "low_confidence_rule_count": 0,
            "review_priority_counts": {"medium_review_priority": 1},
            "rule_scope_counts": {"cross_field": 1},
        },
    )
    json_path = tmp_path / "cross_field_result.json"
    markdown_path = tmp_path / "cross_field_result.md"
    excel_path = tmp_path / "cross_field_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "cross_field_quality_rules" in json_payload
    assert "quality_review_queue_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Cross-Field Quality Rules" in markdown_content
    assert "# Quality Review Queue Summary" in markdown_content

    workbook = load_workbook(excel_path)
    assert "cross_field_quality_rules" in workbook.sheetnames
    assert "quality_review_queue_summary" in workbook.sheetnames


def test_reports_include_confirmed_outputs_when_present(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(
        override_store,
        "MAPPING_OVERRIDES_PATH",
        tmp_path / "mapping_overrides.csv",
    )
    monkeypatch.setattr(
        override_store,
        "STG_OVERRIDES_PATH",
        tmp_path / "stg_overrides.csv",
    )
    monkeypatch.setattr(
        override_store,
        "REVIEW_SESSIONS_DIR",
        tmp_path / "review_sessions",
    )

    override_store.save_mapping_review_records(
        [
            MappingReviewRecord(
                table_name="Sales Order Header",
                field_name="Order__ID",
                original_recommended_standard_code="transaction_id",
                final_standard_code="audit_log_id",
                review_action="edit",
                reviewer_note="report test mapping edit",
                reviewed_at="2026-05-01T10:00:00",
                source="test",
            )
        ]
    )
    override_store.save_stg_review_records(
        [
            StgReviewRecord(
                source_table_name="ods_customer_snapshot",
                source_field_name="snapshot_dt",
                original_recommended_stg_field_name="snapshot_date",
                final_stg_field_name="snapshot_business_date",
                original_recommended_data_type="date",
                final_data_type="timestamp",
                review_action="edit",
                reviewer_note="report test stg edit",
                reviewed_at="2026-05-01T10:00:00",
                source="test",
            )
        ]
    )

    result = run_p0_plus_mapping_plus_stg_with_review_from_file(str(SAMPLE_METADATA_PATH))
    json_path = tmp_path / "confirmed_workflow_result.json"
    markdown_path = tmp_path / "confirmed_workflow_result.md"
    excel_path = tmp_path / "confirmed_workflow_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "confirmed_mapping_results" in json_payload
    assert "confirmed_stg_suggestions" in json_payload
    assert "review_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Confirmed Mapping Results" in markdown_content
    assert "# Confirmed STG Suggestions" in markdown_content

    workbook = load_workbook(excel_path)
    assert "confirmed_standard_mapping" in workbook.sheetnames
    assert "confirmed_stg_fields" in workbook.sheetnames
    assert "review_summary" in workbook.sheetnames


def test_reports_include_confirmed_quality_and_export_results(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="confirmed quality test",
        confirmed_quality_rules=[
            ConfirmedQualityRule(
                source_table_name="sales_order",
                source_field_name="order_id",
                recommended_field_name="order_id",
                rule_type="not_null",
                rule_expression="not_null",
                severity="high",
                priority="P1",
                confirmation_source="override_accept",
                reason="identifier",
            )
        ],
        quality_rule_review_summary={
            "accepted_count": 1,
            "rejected_count": 0,
            "edited_count": 0,
            "manual_review_count": 0,
            "total_reviewed_count": 1,
            "confirmed_count": 1,
        },
        rule_export_results=[
            RuleExportResult(
                export_format="custom_json",
                output_path=str(tmp_path / "rules.json"),
                rule_count=1,
                status="success",
            )
        ],
    )
    json_path = tmp_path / "confirmed_quality_result.json"
    markdown_path = tmp_path / "confirmed_quality_result.md"
    excel_path = tmp_path / "confirmed_quality_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "confirmed_quality_rules" in json_payload
    assert "quality_rule_review_summary" in json_payload
    assert "rule_export_results" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Confirmed Quality Rules" in markdown_content
    assert "# Rule Export Results" in markdown_content

    workbook = load_workbook(excel_path)
    assert "confirmed_quality_rules" in workbook.sheetnames
    assert "quality_rule_review_summary" in workbook.sheetnames
    assert "rule_export_results" in workbook.sheetnames


def test_reports_include_execution_package_outputs(tmp_path: Path) -> None:
    package = ExecutionPackageBuilder().build_package(
        [
            ConfirmedQualityRule(
                source_table_name="sales_order",
                source_field_name="order_id",
                recommended_field_name="order_id",
                rule_type="not_null",
                rule_expression="not_null",
                severity="high",
                priority="P1",
                confirmation_source="override_accept",
                reason="identifier",
            )
        ],
        profile_name="report_package_profile",
    )
    result = WorkflowResult(
        status="success",
        message="execution package test",
        execution_ready_package=package,
        execution_package_summary=ExecutionPackageBuilder.summarize_package(package),
        execution_package_export_results=[
            ExecutionPackageExportResult(
                export_format="package_manifest",
                output_path=str(tmp_path / "manifest.json"),
                package_id=package.package_id,
                rule_count=package.rule_count,
                status="success",
            )
        ],
    )
    json_path = tmp_path / "execution_package_result.json"
    markdown_path = tmp_path / "execution_package_result.md"
    excel_path = tmp_path / "execution_package_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "execution_ready_package" in json_payload
    assert "execution_package_export_results" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Execution-Ready Governance Package" in markdown_content
    assert "# Execution Package Export Results" in markdown_content

    workbook = load_workbook(excel_path)
    assert "execution_ready_rules" in workbook.sheetnames
    assert "execution_package_summary" in workbook.sheetnames
    assert "execution_package_exports" in workbook.sheetnames


def test_reports_include_readiness_and_remediation_outputs(tmp_path: Path) -> None:
    readiness_scores = [
        ReadinessScore(
            object_type="table",
            object_name="sales_order",
            overall_score=0.72,
            readiness_level="partially_ready",
            dimension_scores={"metadata_readiness": 0.8},
        )
    ]
    governance_gaps = [
        GovernanceGap(
            object_type="table",
            object_name="sales_order",
            gap_type="standard_mapping_gap",
            category="mapping",
            severity="medium",
            source_signals=["standard_mapping_low_confidence"],
            reason="Mapping requires review.",
            suggested_owner_role="business_data_steward",
        )
    ]
    remediation_actions = [
        RemediationAction(
            object_type="table",
            object_name="sales_order",
            gap_type="standard_mapping_gap",
            action="Review and confirm standard mappings",
            owner_role="business_data_steward",
            priority="key_tracking",
            expected_output="confirmed mappings",
        )
    ]
    work_package = GovernanceWorkPackage(
        package_name="report_work_package",
        readiness_scores=readiness_scores,
        governance_gaps=governance_gaps,
        remediation_actions=remediation_actions,
        summary="Governance work package report test.",
    )
    result = WorkflowResult(
        status="success",
        message="readiness report test",
        readiness_scores=readiness_scores,
        governance_gaps=governance_gaps,
        remediation_actions=remediation_actions,
        governance_work_package=work_package,
        readiness_summary={"overall_score": 0.72, "gap_count": 1},
    )
    json_path = tmp_path / "readiness_result.json"
    markdown_path = tmp_path / "readiness_result.md"
    excel_path = tmp_path / "readiness_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "readiness_scores" in json_payload
    assert "governance_gaps" in json_payload
    assert "remediation_actions" in json_payload
    assert "governance_work_package" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Governance Readiness Assessment" in markdown_content
    assert "# Governance Gaps" in markdown_content
    assert "# Remediation Plan" in markdown_content
    assert "# Governance Work Package" in markdown_content

    workbook = load_workbook(excel_path)
    assert "readiness_scores" in workbook.sheetnames
    assert "governance_gaps" in workbook.sheetnames
    assert "remediation_actions" in workbook.sheetnames
    assert "governance_work_package_summary" in workbook.sheetnames


def test_reports_include_backlog_outputs(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="backlog report test",
        governance_backlog_items=[
            GovernanceBacklogItem(
                backlog_id="backlog_report_1",
                object_type="table",
                object_name="sales_order",
                gap_type="standard_mapping_gap",
                category="mapping",
                action="Review and confirm standard mappings",
                owner_role="business_data_steward",
                priority="key_tracking",
                status="proposed",
                urgency_score=2,
            )
        ],
        backlog_summary=BacklogSummary(
            total_items=1,
            by_status={"proposed": 1},
            by_priority={"key_tracking": 1},
            by_owner_role={"business_data_steward": 1},
            by_gap_type={"standard_mapping_gap": 1},
            summary="Backlog report test.",
        ),
    )
    json_path = tmp_path / "backlog_result.json"
    markdown_path = tmp_path / "backlog_result.md"
    excel_path = tmp_path / "backlog_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "governance_backlog_items" in json_payload
    assert "backlog_summary" in json_payload
    assert json_payload["backlog_summary"]["total_items"] == 1

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Governance Backlog" in markdown_content
    assert "# Backlog Summary" in markdown_content

    workbook = load_workbook(excel_path)
    assert "governance_backlog_items" in workbook.sheetnames
    assert "backlog_summary" in workbook.sheetnames


def test_reports_include_portfolio_outputs(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="portfolio report test",
        backlog_sla_statuses=[
            BacklogSlaStatus(
                backlog_id="backlog_report_1",
                due_date="2026-05-18",
                age_days=19,
                overdue_days=2,
                is_overdue=True,
                sla_status="overdue",
            )
        ],
        governance_portfolio_summary=GovernancePortfolioSummary(
            total_items=1,
            by_status={"proposed": 1},
            by_priority={"key_tracking": 1},
            by_owner_role={"business_data_steward": 1},
            by_gap_type={"standard_mapping_gap": 1},
            readiness_distribution={"partially_ready": 1},
            overdue_count=1,
            blocked_count=0,
            owner_workload={
                "business_data_steward": {
                    "total": 1,
                    "open": 1,
                    "blocked": 0,
                    "completed": 0,
                    "overdue": 1,
                }
            },
            summary="Portfolio report test.",
        ),
        progress_snapshot=ProgressSnapshot(
            snapshot_id="snapshot_report_1",
            generated_at="2026-05-20T00:00:00",
            total_backlog_items=1,
            completed_count=0,
            blocked_count=0,
            overdue_count=1,
            avg_readiness_score=0.7,
        ),
    )
    json_path = tmp_path / "portfolio_result.json"
    markdown_path = tmp_path / "portfolio_result.md"
    excel_path = tmp_path / "portfolio_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "backlog_sla_statuses" in json_payload
    assert "governance_portfolio_summary" in json_payload
    assert "progress_snapshot" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Backlog SLA Status" in markdown_content
    assert "# Governance Portfolio Summary" in markdown_content
    assert "# Progress Snapshot" in markdown_content

    workbook = load_workbook(excel_path)
    assert "backlog_sla_statuses" in workbook.sheetnames
    assert "governance_portfolio_summary" in workbook.sheetnames
    assert "progress_snapshot" in workbook.sheetnames


def test_reports_include_delivery_package_outputs(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="delivery report test",
        confirmation_workbook_results=[
            ConfirmationWorkbookResult(
                workbook_type="mapping_confirmation",
                output_path=str(tmp_path / "mapping.xlsx"),
                row_count=1,
                status="success",
            )
        ],
        governance_delivery_manifest=GovernanceDeliveryManifest(
            package_name="delivery_report",
            included_artifacts=[
                {
                    "artifact_type": "mapping_confirmation",
                    "path": str(tmp_path / "mapping.xlsx"),
                    "row_count": 1,
                    "status": "success",
                }
            ],
        ),
        governance_delivery_package_result=GovernanceDeliveryPackageResult(
            package_name="delivery_report",
            output_dir=str(tmp_path),
            generated_files={"package_manifest": str(tmp_path / "manifest.json")},
            status="success",
        ),
    )
    json_path = tmp_path / "delivery_result.json"
    markdown_path = tmp_path / "delivery_result.md"
    excel_path = tmp_path / "delivery_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "confirmation_workbook_results" in json_payload
    assert "governance_delivery_manifest" in json_payload
    assert "governance_delivery_package_result" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Governance Delivery Package" in markdown_content

    workbook = load_workbook(excel_path)
    assert "confirmation_workbooks" in workbook.sheetnames
    assert "delivery_manifest" in workbook.sheetnames
    assert "delivery_package" in workbook.sheetnames


def test_reports_include_batch_and_incremental_outputs(tmp_path: Path) -> None:
    result = WorkflowResult(
        status="success",
        message="batch report test",
        batch_group_results=[
            BatchGroupResult(
                group_name="crm",
                file_count=2,
                table_count=3,
                status="success",
            )
        ],
        incremental_diff_items=[
            IncrementalDiffItem(
                object_type="table",
                object_name="crm.customer",
                group_name="crm",
                diff_type="changed",
            )
        ],
        incremental_diff_summary=IncrementalDiffSummary(
            total_objects=1,
            new_count=0,
            changed_count=1,
            unchanged_count=0,
            removed_count=0,
            pending_review_count=0,
        ),
        rerun_scope_summary={"rerun_object_count": 1, "changed_only": True},
    )
    json_path = tmp_path / "batch_result.json"
    markdown_path = tmp_path / "batch_result.md"
    excel_path = tmp_path / "batch_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "batch_group_results" in json_payload
    assert "incremental_diff_items" in json_payload
    assert "rerun_scope_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Batch Processing Summary" in markdown_content
    assert "# Incremental Diff Summary" in markdown_content
    assert "# Rerun Scope Summary" in markdown_content

    workbook = load_workbook(excel_path)
    assert "batch_group_results" in workbook.sheetnames
    assert "incremental_diff_items" in workbook.sheetnames
    assert "incremental_diff_summary" in workbook.sheetnames


def test_reports_include_workbook_import_and_roundtrip_outputs(tmp_path: Path) -> None:
    import_summary = WorkbookImportSummary(
        workbook_type="mapping_confirmation",
        total_rows=1,
        imported_count=1,
        skipped_count=0,
        invalid_count=0,
        accepted_count=1,
        rejected_count=0,
        edited_count=0,
        manual_review_count=0,
    )
    result = WorkflowResult(
        status="success",
        message="roundtrip report test",
        workbook_import_summaries=[import_summary],
        roundtrip_results=[
            ConfirmationRoundTripResult(
                workbook_type="mapping_confirmation",
                import_summary=import_summary,
                generated_review_records_count=1,
                generated_override_updates_count=1,
                changed_object_keys=["customer.customer_id"],
                status="success",
            )
        ],
        roundtrip_changed_objects_summary={"changed_object_count": 1},
    )
    json_path = tmp_path / "roundtrip_result.json"
    markdown_path = tmp_path / "roundtrip_result.md"
    excel_path = tmp_path / "roundtrip_result.xlsx"

    export_workflow_result_to_json(result, str(json_path))
    export_workflow_result_to_markdown(result, str(markdown_path))
    export_workflow_result_to_excel(result, str(excel_path))

    json_payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert "workbook_import_summaries" in json_payload
    assert "roundtrip_results" in json_payload
    assert "roundtrip_changed_objects_summary" in json_payload

    markdown_content = markdown_path.read_text(encoding="utf-8")
    assert "# Workbook Import Summary" in markdown_content
    assert "# Confirmation Round-Trip Results" in markdown_content

    workbook = load_workbook(excel_path)
    assert "workbook_import_summary" in workbook.sheetnames
    assert "roundtrip_results" in workbook.sheetnames
